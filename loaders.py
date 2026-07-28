import json
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook



# PDF解析
def parse_pdf(path):

    reader = PdfReader(path)

    text = ""

    for page in reader.pages:
        text += page.extract_text()

    return text



# Word解析
def parse_word(path):

    doc = Document(path)

    text = ""

    for para in doc.paragraphs:
        text += para.text + "\n"

    return text



# Excel解析

def parse_excel(path):

    wb=load_workbook(path)


    text=""


    for sheet in wb:


        for row in sheet.iter_rows():


            line=[]


            for cell in row:


                if cell.value is not None:

                    line.append(
                        str(cell.value)
                    )


            if line:

                text+=" ".join(line)+"\n"


    return text


# TXT解析

def parse_txt(path):

    with open(path,"r",encoding="utf-8") as f:
        return f.read()



# JSON解析

def parse_json(path):

    with open(path,"r",encoding="utf-8") as f:

        data=json.load(f)


    return json.dumps(
        data,
        ensure_ascii=False,
        indent=2
    )




if __name__=="__main__":

    print(parse_pdf("data/test.pdf"))
